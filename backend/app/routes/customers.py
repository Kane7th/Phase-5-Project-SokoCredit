from flask import Blueprint, request, jsonify, Response, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.extensions import db
from app.models.customer import Customer
from utils.decorators import role_required
from sqlalchemy import or_
import csv
import io
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from app.models.notification import Notification
from app.models.user import User
from utils.sms_service import send_sms

customers_bp = Blueprint('customers', __name__)

<<<<<<< HEAD
UPLOAD_FOLDER = "uploads/customers"

# Helpers
=======
@customers_bp.route("/by-user/<int:user_id>", methods=["GET"])
@jwt_required()
@role_required(["admin", "lender", "customer"])
def get_customer_by_user(user_id):
    current_user_id, role = extract_identity()
    if current_user_id != user_id and role not in ["admin", "lender"]:
        return jsonify({"msg": "Not authorized to view this customer"}), 403

    customer = Customer.query.filter_by(customer_user_id=user_id).first()
    if not customer:
        # Try to find customer created by this user as fallback
        customer = Customer.query.filter_by(created_by=user_id).first()
        if not customer:
            return jsonify({"msg": "Customer profile not found"}), 404

    return jsonify(format_customer(customer)), 200

>>>>>>> 99e5205a91df05ed5dd41c5fe54ea99ce22f6ce0
def extract_identity():
    identity = get_jwt_identity()
    user_id_str, role = identity.split(":")
    return int(user_id_str.replace("user_", "")), role

def is_authorized(customer, user_id, role):
    if role == "customer":
        return customer.customer_user_id == user_id
    return role in ["admin", "lender"] or customer.created_by == user_id

def format_customer(customer):
    return {
        "id": customer.id,
        "full_name": customer.full_name,
        "phone": customer.phone,
        "business_name": customer.business_name,
        "location": customer.location,
        "documents": customer.documents
    }

def generate_csv(customers):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Full Name", "Phone", "Business", "Location", "Documents"])
    for c in customers:
        writer.writerow([
            c.id, c.full_name, c.phone, c.business_name, c.location, str(c.documents)
        ])
    output.seek(0)
    return output.read()

def generate_excel(customers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Full Name", "Phone", "Business", "Location", "Documents"])
    for c in customers:
        ws.append([
            c.id, c.full_name, c.phone, c.business_name, c.location, str(c.documents)
        ])
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

# Check if file is allowed
def allowed_file(filename):
    allowed_exts = {"pdf", "jpg", "jpeg", "png"}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_exts

# Routes
@customers_bp.route("/", methods=["POST"])
@jwt_required()
@role_required(["admin", "lender", "customer"])
def create_customer():
    data = request.get_json()
    user_id, role = extract_identity()

<<<<<<< HEAD
    user = User.query.get(user_id)
    if not data.get("phone") and not (user and user.phone):
        return jsonify({"msg": "Phone is required"}), 400

    full_name = data.get("full_name")
    if not full_name and user:
        full_name = f"{user.first_name or ''} {user.middle_name or ''} {user.last_name or ''}".strip()

    if not full_name:
        return jsonify({"msg": "Full name is required"}), 400

    phone = data.get("phone") or user.phone
    email = data.get("email") or user.email

    if Customer.query.filter_by(phone=phone).first():
=======
    if not data.get("full_name") or not data.get("phone"):
        return jsonify({"msg": "Full name and phone are required"}), 400

    # Ensure unique phone
    if Customer.query.filter_by(phone=data["phone"]).first():
>>>>>>> 99e5205a91df05ed5dd41c5fe54ea99ce22f6ce0
        return jsonify({"msg": "Phone number already exists"}), 409

    if role == "customer":
        if Customer.query.filter_by(customer_user_id=user_id).first():
            return jsonify({"msg": "You already have a profile"}), 409
        customer_user_id = user_id
    else:
        customer_user_id = data.get("customer_user_id")
        if not customer_user_id:
            return jsonify({"msg": "customer_user_id is required for admin/lender"}), 400

<<<<<<< HEAD
    customer = Customer(
        full_name=full_name,
        phone=phone,
        email=email,
        business_name=data.get("business_name"),
        location=data.get("location"),
        documents=data.get("documents", {}),
        created_by=user_id,
        customer_user_id=customer_user_id
    )
=======
    customer_kwargs = {
        "full_name": data["full_name"],
        "phone": data["phone"],
        "email": data.get("email"),
        "business_name": data.get("business_name"),
        "location": data.get("location"),
        "documents": data.get("documents", {}),
        "created_by": user_id,
        "customer_user_id": customer_user_id
    }

    customer = Customer(**customer_kwargs)
>>>>>>> 99e5205a91df05ed5dd41c5fe54ea99ce22f6ce0

    db.session.add(customer)
    db.session.commit()

    Notification.create_notification(user_id=user_id, message=f"Customer profile created for {full_name}.")
    if user and user.phone:
        send_sms(user.phone, f"SokoCredit: Your customer profile '{full_name}' has been created.")

    print(f"[AUDIT LOG] User {user_id} ({role}) created customer {customer.id} at {datetime.utcnow().isoformat()}.")
    return jsonify({"msg": "Customer created", "id": customer.id}), 201

<<<<<<< HEAD
=======
@customers_bp.route("/", methods=["GET"])
@jwt_required()
@role_required(["admin", "lender"])
def list_customers():
    # Validate allowed query params
    ALLOWED_FILTERS = {
        "page", "per_page", "search", "created_by",
        "business_name", "has_documents", "location", "format"
    }
    for key in request.args:
        if key not in ALLOWED_FILTERS:
            return jsonify({"msg": f"Invalid filter: '{key}' is not a valid filter field"}), 400

    # Extract query parameters
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    search = request.args.get("search", "", type=str)
    created_by = request.args.get("created_by", type=int)
    business_name = request.args.get("business_name", type=str)
    has_documents = request.args.get("has_documents")
    location = request.args.get("location")
    format_type = request.args.get("format", "").lower()

    query = Customer.query

    # Filtering
    if location and location.strip():
        query = query.filter(Customer.location == location)

    if has_documents is not None:
        if has_documents.lower() == "true":
            query = query.filter(Customer.documents != {})
        elif has_documents.lower() == "false":
            query = query.filter(Customer.documents == {})

    if created_by is not None:
        query = query.filter(Customer.created_by == created_by)

    if business_name and business_name.strip():
        query = query.filter(Customer.business_name == business_name)

    if search:
        search_term = f"%{search}%"
        query = query.filter(or_(
            Customer.full_name.ilike(search_term),
            Customer.phone.ilike(search_term),
            Customer.business_name.ilike(search_term),
            Customer.location.ilike(search_term)
        ))

    # CSV or Excel export
    if format_type == "csv":
        customers = query.all()
        csv_data = generate_csv(customers)
        return Response(csv_data, mimetype="text/csv", headers={
            "Content-Disposition": "attachment; filename=customers.csv"
        })

    if format_type == "excel":
        customers = query.all()
        xlsx_data = generate_excel(customers)
        return Response(xlsx_data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
            "Content-Disposition": "attachment; filename=customers.xlsx"
        })
    
    # audit log for listing customers
    print(f"[AUDIT LOG] User {get_jwt_identity()} listed customers at {datetime.utcnow().isoformat()}.")

    # Paginate + return JSON
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({
        "customers": [format_customer(c) for c in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
        "pages": pagination.pages
    })

@customers_bp.route("/<int:customer_id>", methods=["GET"])
@jwt_required()
@role_required(["admin", "lender", "customer"])
def get_customer(customer_id):
    user_id, role = extract_identity()
    customer = Customer.query.get_or_404(customer_id)

    if not is_authorized(customer, user_id, role):
        return jsonify({"msg": "Not authorized to view this customer"}), 403
    
    print(f"[AUDIT LOG] User {user_id} ({role}) viewed customer {customer.id} at {datetime.utcnow().isoformat()}.")

    return jsonify(format_customer(customer)), 200

from flask_cors import cross_origin

@customers_bp.route("/my_customers", methods=["GET", "OPTIONS"])
@cross_origin()
@jwt_required()
@role_required(["admin", "lender"])
def get_my_customers():
    if request.method == 'OPTIONS':
        response = jsonify({})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Methods", "GET, OPTIONS")
        response.headers.add("Access-Control-Allow-Headers", "Authorization, Content-Type")
        return response
        
    identity = get_jwt_identity()
    print(f"DEBUG: JWT identity received: {identity}")
    # Try to parse user_id from identity string
    try:
        # If identity is string like "15:lender", split by ":" and convert first part to int
        if isinstance(identity, str) and ":" in identity:
            user_id_str, role = identity.split(":")
            if user_id_str.startswith("user_"):
                user_id_str = user_id_str[len("user_"):]
            user_id = int(user_id_str)
        elif isinstance(identity, int):
            user_id = identity
        else:
            raise ValueError("Unexpected identity format")
    except Exception as e:
        print(f"Error parsing user_id from identity: {e}")
        return jsonify({"msg": "Invalid user identity format"}), 400

    # Check if lender_id query parameter is provided
    lender_id = request.args.get('lender_id', type=int)
    if lender_id:
        customers = Customer.query.filter_by(lender_id=lender_id).all()
        print(f"[AUDIT LOG] User {user_id} listed customers for lender {lender_id} at {datetime.utcnow().isoformat()}.")
        print(f"DEBUG: Found {len(customers)} customers for lender_id {lender_id}")
    else:
        customers = Customer.query.filter_by(lender_id=user_id).all()
        print(f"[AUDIT LOG] User {user_id} listed their own customers at {datetime.utcnow().isoformat()}.")
        print(f"DEBUG: Found {len(customers)} customers for lender_id {user_id}")

    return jsonify([c.to_dict() for c in customers])

>>>>>>> 99e5205a91df05ed5dd41c5fe54ea99ce22f6ce0
@customers_bp.route("/<int:customer_id>", methods=["PATCH"])
@jwt_required()
@role_required(["admin", "lender", "customer"])
def patch_customer(customer_id):
    data = request.get_json()
    user_id, role = extract_identity()
    customer = Customer.query.get_or_404(customer_id)

    if not is_authorized(customer, user_id, role):
        return jsonify({"msg": "Not authorized to update this customer"}), 403

    allowed_fields = ["full_name", "phone", "business_name", "location"]
    if role in ["admin", "lender"]:
        allowed_fields.append("documents")

    for field in allowed_fields:
        if field in data:
            setattr(customer, field, data[field])

    if "phone" in data:
        existing = Customer.query.filter(Customer.phone == data["phone"], Customer.id != customer.id).first()
        if existing:
            return jsonify({"msg": "Phone number already exists for another customer"}), 409

    db.session.commit()

    Notification.create_notification(user_id=user_id, message=f"Customer #{customer.id} updated.")
    if role == "customer" and customer.phone:
        send_sms(customer.phone, f"SokoCredit: Your profile has been updated.")

    print(f"[AUDIT LOG] User {user_id} ({role}) updated customer {customer.id} at {datetime.utcnow().isoformat()}.")
    return jsonify({"msg": "Customer updated", "customer": format_customer(customer)}), 200

@customers_bp.route("/<int:customer_id>", methods=["DELETE"])
@jwt_required()
@role_required(["admin"])
def delete_customer(customer_id):
    user_id, role = extract_identity()
    customer = Customer.query.get_or_404(customer_id)

    if not is_authorized(customer, user_id, role):
        return jsonify({"msg": "Not authorized to delete this customer"}), 403

    Notification.create_notification(user_id=user_id, message=f"Customer #{customer.id} deleted.")
    if customer.phone:
        send_sms(customer.phone, f"SokoCredit: Your profile has been deleted.")

    print(f"[AUDIT LOG] User {user_id} ({role}) deleted customer {customer.id} at {datetime.utcnow().isoformat()}.")
    db.session.delete(customer)
    db.session.commit()
    return '', 204

@customers_bp.route("/<int:customer_id>/upload", methods=["POST"])
@jwt_required()
@role_required(["admin", "lender", "customer"])
def upload_customer_document(customer_id):
    user_id, role = extract_identity()
    customer = Customer.query.get_or_404(customer_id)

    if not is_authorized(customer, user_id, role):
        return jsonify({"msg": "Not authorized to upload files for this customer"}), 403

    if 'file' not in request.files:
        return jsonify({"msg": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"msg": "No selected file"}), 400
    if not allowed_file(file.filename):
        return jsonify({"msg": "File type not allowed"}), 400

    doc_type = request.form.get("doc_type", "unknown")
    filename = secure_filename(file.filename)
    customer_dir = os.path.join(UPLOAD_FOLDER, str(customer_id))
    os.makedirs(customer_dir, exist_ok=True)

    file_path = os.path.join(customer_dir, f"{doc_type}_{filename}")
    file.save(file_path)

    customer.documents[doc_type] = file_path
    db.session.commit()

    Notification.create_notification(user_id=user_id, message=f"Document '{doc_type}' uploaded for customer #{customer.id}.")
    if customer.phone:
        send_sms(customer.phone, f"SokoCredit: Your document '{doc_type}' has been uploaded.")

<<<<<<< HEAD
    print(f"[AUDIT LOG] User {user_id} ({role}) uploaded '{doc_type}' for customer {customer.id} at {datetime.utcnow().isoformat()}.")
    return jsonify({"msg": "File uploaded", "path": file_path, "documents": customer.documents}), 200
=======
    return jsonify({"msg": "File uploaded", "path": file_path, "documents": customer.documents}), 200
>>>>>>> 99e5205a91df05ed5dd41c5fe54ea99ce22f6ce0
